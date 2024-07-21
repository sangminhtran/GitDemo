import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

file_path = "/Users/sangtran/Downloads/download.xlsx"
fruit_name = "Apple"
newValue = "999"

def update_excel_data(file_path, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column+1):  # tìm cột price và lấy vị trí của cột
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):  # tìm item Apple và lấy vị trí hàng
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)



driver = webdriver.Chrome()
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()

#edit the excel file with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

#upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 10)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

# smart dynamic Xpath (giả sử sau dev có add thêm cột vào bảng thì mình ko cần update code)
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
#từ element child truy cập ngược lên parent rồi từ parent xuống child khác thì cú pháp như dưới
#cách này để có thể truy vấn data ngang hàng trong bảng, chỉ cần thay text check value khác
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actual_price == newValue

driver.get_screenshot_as_file("upload.png")

